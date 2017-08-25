# -*- coding: utf-8 -*-
# ****************************************************************
# testframe.py
# Author     : LUCKY
# Version    : 1.0
# Date       : 2016-05-04
# Description: 测试组装，用例执行入口
# ****************************************************************
from openpyxl import load_workbook
import requests


class create_excel():
    # 初始化类
    def __init__(self, file, resultfile, argbegin, titleindex, casebegin, url):
        # 文件路径
        self.file = file
        self.resultfile = resultfile
        # excel
        self.wb = load_workbook(file)
        # 参数开始位置
        self.argbegin = argbegin
        # 用例标题索引
        self.titleindex = titleindex
        # 用例开始索引
        self.casebegin = casebegin
        # url
        self.url = url
        # 默认使用第一个sheet
        self.ws = self.wb.active
        # 用例个数
        self.casenum = self.ws.max_row - 2
        # 接口信息
        self.names = self.read_data(1, 2).split(',')
        # 接口名称
        self.name = self.names[0]
        # 请求方式
        self.httptype = self.names[1]
        # 参数个数
        self.argcount = int(self.names[2])
        # 预期结果索引
        self.expr = self.argcount + self.argbegin
        # 实际结果索引
        self.realr = self.expr + 1
        # TestResult
        self.realt = self.realr + 1

    # 读取excel单元格
    def read_data(self, row, column):
        sValue = self.ws.cell(row=row, column=column).value
        # 去除'.0'
        # if sValue[-2:] == '.0':
        #     sValue = sValue[0:-2]
        return sValue

    # 写入excel数据
    def write_data(self, row, column, sdata):
        self.ws.cell(row=row, column=column).value = sdata
        self.wb.save(filename=self.resultfile)  # resultfile

    # 生成请求数据,根据用例索引
    def do_data(self, caseid):
        # ddicts = []
        # for j in range(2, self.ws.max_row - 1):
        ddict = {}
        for i in range(3, self.argcount + 3):
            ddict[self.read_data(2, i)] = self.read_data(2 + caseid, i)
        # ddicts.append(ddict)
        return ddict

    def assertres(self):
        # 遍历执行case
        case_nt = 0
        case_ok = 0
        case_ng = 0
        for caseid in range(0, self.casenum):
            # 检查是否执行该Case
            realrr = request(self.url, self.name, self.do_data(caseid), self.httptype)
            if self.read_data(self.casebegin + caseid, 2) == 'N':
                self.write_data(self.casebegin + caseid, self.realt, 'NT')
                case_nt = case_nt + 1
            # 判断是否和预期相同
            elif realrr == self.read_data(self.casebegin + caseid, self.expr):
                self.write_data(self.casebegin + caseid, self.realr, realrr)
                self.write_data(self.casebegin + caseid, self.realt, 'OK')
                case_ok = case_ok + 1
            else:
                self.write_data(self.casebegin + caseid, self.realr, realrr)
                self.write_data(self.casebegin + caseid, self.realt, 'NG')
                case_ng = case_ng + 1
        msg = '未测试:' + str(case_nt) + ',通过:' + str(case_ok) + ',未通过:' + str(case_ng)
        return (case_nt, case_ok, case_ng, msg)


# 执行调用
def request(url, name, data, httptype):
    if httptype == 'get':
        r = requests.get(url + name, params=data)
    else:
        r = requests.post(url + name, data=data)
    return r.text
