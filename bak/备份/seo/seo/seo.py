from openpyxl import load_workbook
from bottle import template

# 读取html模板信息
file2 = r'E:\script\新框架\seotem.html'
f = open(file2, 'r', encoding='utf-8')
template_demo = f.read()

# 读取用户信息
file = r'C:\Users\Administrator\Desktop\seo.xlsx'
wb = load_workbook(file)

ws = wb.get_sheet_by_name('1')
row = len(ws.rows)
for i in range(2, row + 1):  # i + 1
    zb = ws.cell(row=i, column=1).value
    print(zb)
    title = '%s的直播间_%s的视频直播-火猫直播' % (zb, zb)
    description = '火猫直播为您提供主播%s的精彩直播内容,了解%s最新直播讯息，关注%s直播的最新动态,快来火猫直播订阅%s的视频直播间吧!' % (zb, zb, zb, zb)
    keywords = '%s直播、%s视频直播' % (zb, zb)
    roomname = '%s的直播间' % zb
    zbname = zb
    html = template(template_demo, title=title, roomname=roomname, zbname=zbname, description=description, keywords=keywords)
    file3 = 'C:\\Users\\Administrator\\Desktop\\test\\' + str(i) + '.html'
    f2 = open(file3, 'w', encoding='utf-8')
    f2.write(html)
    f2.close()


# ws = wb.get_sheet_by_name('2')
# row = len(ws.rows)
# for i in range(2, row + 1):  # i + 1
#     zb = ws.cell(row=i, column=1).value
#     print(zb)
#     title = '%s英雄联盟直播_%sLOL视频直播-火猫直播' % (zb, zb)
#     description = '%sLOL的直播间，%s与您分享英雄联盟游戏乐趣。实时观看更多%s在游戏英雄联盟中的精彩表现，就在火猫直播%s英雄联盟的直播间!' % (zb, zb, zb, zb)
#     keywords = '%s直播、%s英雄联盟直播、%sLOL视频' % (zb, zb, zb)
#     roomname = '%s的直播间' % zb
#     zbname = zb
#     html = template(template_demo, title=title, roomname=roomname, zbname=zbname, description=description, keywords=keywords)
#     file3 = 'C:\\Users\\Administrator\\Desktop\\test\\' + str(i) + 'lol.html'
#     f2 = open(file3, 'w', encoding='utf-8')
#     f2.write(html)
#     f2.close()
